"""
Xero ↔ Flash Purchases Reconciliation

Accepts a Xero "Account Transactions" XLSX export, parses it,
matches rows against Flash invoices for the same period, and
returns a four-section reconciliation report.

Matching is rule-based first (exact → near-match), with an
optional LLM pass for remaining unmatched items.
"""
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional
import io
import re
import logging
import json

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from pydantic import BaseModel

from database import get_db
from models.user import User
from models.invoice import Invoice, InvoiceStatus
from models.supplier import Supplier
from auth.jwt import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter()


# ============ Pydantic Response Models ============

class ReconcileMatchedItem(BaseModel):
    date: str
    supplier: str
    ref: str
    amount: str  # formatted £
    flash_id: int | None = None
    match_source: str = "rule"  # "rule" or "ai"

class ReconcileDiscrepancyItem(BaseModel):
    date: str
    supplier: str
    ref: str
    flash_amount: str
    xero_amount: str
    difference: str
    differs: list[str]  # which fields differ: "amount", "date", "ref"
    flash_id: int | None = None
    flash_date: str | None = None
    xero_date: str | None = None
    flash_ref: str | None = None
    xero_ref: str | None = None
    flash_supplier: str | None = None
    xero_description: str | None = None
    match_source: str = "rule"  # "rule" or "ai"
    amount_insight: str | None = None  # explanation for amount discrepancy

class ReconcileUnmatchedFlash(BaseModel):
    date: str
    supplier: str
    ref: str
    net_stock: str
    flash_id: int

class ReconcileUnmatchedXero(BaseModel):
    date: str
    description: str
    ref: str
    net: str
    is_expected_external: bool = False

class ReconcileResponse(BaseModel):
    period_start: str
    period_end: str
    flash_total: str
    xero_total: str
    difference: str
    matched_count: int
    discrepancy_count: int
    flash_only_count: int
    xero_only_count: int
    non_stock_excluded_count: int
    non_stock_excluded_total: str
    matched: list[ReconcileMatchedItem]
    discrepancies: list[ReconcileDiscrepancyItem]
    flash_only: list[ReconcileUnmatchedFlash]
    xero_only: list[ReconcileUnmatchedXero]
    llm_matches_attempted: bool = False


# ============ Supplier alias map for Xero description → Flash supplier ============

# Xero descriptions often look like "Supplier Name - food" or "Supplier - food crn"
# This maps known Xero description variants to the Flash supplier name
SUPPLIER_ALIASES = {
    "j hall & son (bakers) ltd": "Halls",
    "j hall & son bakers ltd": "Halls",
    "halls": "Halls",
    "r & d walker ltd": "R&D Walker",
    "r&d walker ltd": "R&D Walker",
    "r&d walker": "R&D Walker",
    "lambournes of stow-on-the-wold": "Lambournes",
    "lambournes": "Lambournes",
    "bramleys": "Bramleys",
    "cotswold coffee": "Cotswold Coffee",
    "brakes": "Brakes",
    "direct seafoods": "Direct Seafoods",
}

# Xero-only entries that are expected (not from Flash) — collapsible sub-section
EXPECTED_EXTERNAL_PATTERNS = [
    "tesco",
    "revenue jv",
    "journal",
    "internal",
    "petty",
]


# ============ Helpers ============

def normalise_ref(ref: str | None) -> str:
    """Normalise a reference for comparison: uppercase, strip whitespace/leading #, collapse spaces."""
    if not ref:
        return ""
    r = ref.strip().upper()
    r = r.lstrip("#")
    r = re.sub(r"\s+", " ", r).strip()
    return r


def extract_supplier_from_xero_desc(description: str) -> str:
    """
    Extract supplier name from Xero description.
    Xero format: "Supplier Name - food", "Supplier - food crn", etc.
    """
    desc = description.strip()
    # Strip trailing " - food", " - food crn", " - beverage", etc.
    desc = re.sub(r"\s*-\s*(food|beverage|cleaning|sundry)(\s+crn)?\s*$", "", desc, flags=re.IGNORECASE)
    return desc.strip()


def normalise_supplier(name: str) -> str:
    """Normalise supplier name to lowercase for comparison."""
    return name.strip().lower()


def resolve_supplier(xero_desc: str) -> str:
    """Resolve a Xero description to a canonical supplier name."""
    extracted = extract_supplier_from_xero_desc(xero_desc)
    norm = normalise_supplier(extracted)
    if norm in SUPPLIER_ALIASES:
        return SUPPLIER_ALIASES[norm]
    return extracted


def is_expected_external(xero_desc: str) -> bool:
    """Check if a Xero entry is expected to be outside Flash."""
    desc_lower = xero_desc.lower()
    return any(pat in desc_lower for pat in EXPECTED_EXTERNAL_PATTERNS)


def levenshtein(s1: str, s2: str) -> int:
    """Compute Levenshtein distance between two strings."""
    if len(s1) < len(s2):
        return levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev_row = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = prev_row[j + 1] + 1
            deletions = curr_row[j] + 1
            substitutions = prev_row[j] + (c1 != c2)
            curr_row.append(min(insertions, deletions, substitutions))
        prev_row = curr_row
    return prev_row[-1]


def alphanumeric_only(s: str) -> str:
    """Strip non-alphanumeric characters for fuzzy ref comparison."""
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def amounts_match(a: Decimal, b: Decimal, tolerance: Decimal = Decimal("0.01")) -> bool:
    return abs(a - b) <= tolerance


def fmt_money(d: Decimal) -> str:
    """Format a decimal as £X,XXX.XX"""
    return f"£{d:,.2f}"


def parse_xero_date(val) -> date | None:
    """Parse a date from an openpyxl cell value."""
    from datetime import datetime
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d %b %Y", "%d/%m/%Y", "%Y-%m-%d", "%d %B %Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def parse_xero_amount(val) -> Decimal:
    """Parse a numeric cell to Decimal, returning 0 for None/empty."""
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(val)).quantize(Decimal("0.01"))
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("£", "")
        if not val or val == "-":
            return Decimal("0")
        try:
            return Decimal(val).quantize(Decimal("0.01"))
        except InvalidOperation:
            return Decimal("0")
    return Decimal("0")


# ============ XLSX Parsing ============

def parse_xero_xlsx(file_bytes: bytes) -> tuple[date, date, list[dict]]:
    """
    Parse a Xero Account Transactions XLSX export.

    Returns (period_start, period_end, rows) where each row is:
    {
        "date": date,
        "description": str,
        "ref": str,
        "debit": Decimal,
        "credit": Decimal,
        "net": Decimal,  # debit - credit
    }
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Extract period from header rows (typically row 3)
    period_start = None
    period_end = None
    for row_idx in range(1, 6):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and isinstance(cell_val, str) and "period" in cell_val.lower():
            # "For the period 1 March 2026 to 31 March 2026"
            match = re.search(
                r"(\d{1,2}\s+\w+\s+\d{4})\s+to\s+(\d{1,2}\s+\w+\s+\d{4})",
                cell_val, re.IGNORECASE
            )
            if match:
                from datetime import datetime
                for fmt in ("%d %B %Y", "%d %b %Y"):
                    try:
                        period_start = datetime.strptime(match.group(1), fmt).date()
                        period_end = datetime.strptime(match.group(2), fmt).date()
                        break
                    except ValueError:
                        continue
            break

    if not period_start or not period_end:
        raise ValueError("Could not extract report period from XLSX header rows. Expected 'For the period DD Month YYYY to DD Month YYYY' in rows 1-5.")

    # Find header row (look for "Date" in column A)
    header_row = None
    for row_idx in range(1, 10):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and isinstance(cell_val, str) and cell_val.strip().lower() == "date":
            header_row = row_idx
            break

    if not header_row:
        raise ValueError("Could not find column header row (expected 'Date' in column A within first 10 rows).")

    # Map column headers
    col_map = {}
    for col_idx in range(1, 20):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and isinstance(val, str):
            col_map[val.strip().lower()] = col_idx

    required = ["date", "description", "debit", "credit"]
    for req in required:
        if req not in col_map:
            raise ValueError(f"Missing required column '{req}' in header row {header_row}. Found: {list(col_map.keys())}")

    ref_col = col_map.get("reference", col_map.get("ref"))
    source_col = col_map.get("source")

    # Parse data rows
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=col_map["date"]).value
        parsed_date = parse_xero_date(date_val)
        if not parsed_date:
            continue  # Skip non-data rows (section headers, totals, blanks)

        description = ws.cell(row=row_idx, column=col_map["description"]).value or ""
        ref = ""
        if ref_col:
            ref = ws.cell(row=row_idx, column=ref_col).value or ""
        ref = str(ref).strip()

        source = ""
        if source_col:
            source = str(ws.cell(row=row_idx, column=source_col).value or "").strip()

        debit = parse_xero_amount(ws.cell(row=row_idx, column=col_map["debit"]).value)
        credit = parse_xero_amount(ws.cell(row=row_idx, column=col_map["credit"]).value)
        net = debit - credit  # Credits (credit notes) become negative

        is_credit_note = "credit note" in source.lower() or credit > 0

        rows.append({
            "date": parsed_date,
            "description": str(description).strip(),
            "ref": ref,
            "source": source,
            "debit": debit,
            "credit": credit,
            "net": net,
            "is_credit_note": is_credit_note,
        })

    wb.close()
    return period_start, period_end, rows


def group_xero_by_ref(rows: list[dict]) -> list[dict]:
    """
    Group Xero rows by reference (Xero sometimes splits an invoice across
    stock/non-stock lines with the same ref). Sum the nets.

    Rows with empty refs are NOT grouped — each stays separate.
    """
    grouped = {}
    no_ref = []

    for row in rows:
        norm = normalise_ref(row["ref"])
        if not norm:
            no_ref.append({
                "date": row["date"],
                "description": row["description"],
                "ref": row["ref"],
                "net": row["net"],
                "is_credit_note": row.get("is_credit_note", False),
                "raw_rows": [row],
            })
        else:
            if norm not in grouped:
                grouped[norm] = {
                    "date": row["date"],
                    "description": row["description"],
                    "ref": row["ref"],  # keep original formatting from first row
                    "net": Decimal("0"),
                    "is_credit_note": row.get("is_credit_note", False),
                    "raw_rows": [],
                }
            grouped[norm]["net"] += row["net"]
            grouped[norm]["raw_rows"].append(row)
            # If any row is a credit note, mark the group
            if row.get("is_credit_note"):
                grouped[norm]["is_credit_note"] = True

    return list(grouped.values()) + no_ref


# ============ Flash Invoice Querying ============

async def get_flash_invoices(
    db: AsyncSession,
    kitchen_id: int,
    period_start: date,
    period_end: date,
) -> tuple[list[dict], int, Decimal]:
    """
    Get Flash invoices for the period. Returns:
    - stock_invoices: list of dicts for matching
    - non_stock_count: count of excluded non-stock-only invoices
    - non_stock_total: total of excluded non-stock amounts

    Each stock invoice dict has:
    {
        "id": int,
        "date": date,
        "supplier_name": str,
        "ref": str (invoice_number),
        "net_stock": Decimal,
        "document_type": str,
    }
    """
    from models.line_item import LineItem

    result = await db.execute(
        select(Invoice)
        .where(
            Invoice.kitchen_id == kitchen_id,
            Invoice.status == InvoiceStatus.CONFIRMED,
        )
        .options(selectinload(Invoice.line_items))
        .order_by(Invoice.invoice_date.desc().nullslast())
    )
    all_invoices = result.scalars().all()

    # Get supplier map
    supplier_result = await db.execute(
        select(Supplier).where(Supplier.kitchen_id == kitchen_id)
    )
    suppliers_map = {s.id: s for s in supplier_result.scalars().all()}

    stock_invoices = []
    non_stock_count = 0
    non_stock_total = Decimal("0")

    for inv in all_invoices:
        inv_date = inv.invoice_date or inv.created_at.date()
        if not (period_start <= inv_date <= period_end):
            continue

        # Calculate net_stock (stock items only)
        net_stock = Decimal("0")
        net_non_stock = Decimal("0")
        if inv.line_items:
            for item in inv.line_items:
                item_net = item.amount or Decimal("0")
                if item.is_non_stock:
                    net_non_stock += item_net
                else:
                    net_stock += item_net

        # Credit notes: negate if positive
        if inv.document_type == "credit_note":
            if net_stock > 0:
                net_stock = -net_stock
            if net_non_stock > 0:
                net_non_stock = -net_non_stock

        # Get supplier name
        supplier_name = ""
        if inv.supplier_id and inv.supplier_id in suppliers_map:
            supplier_name = suppliers_map[inv.supplier_id].name
        elif inv.vendor_name:
            supplier_name = inv.vendor_name
        else:
            supplier_name = "Unknown"

        if net_stock == 0 and net_non_stock != 0:
            # Entirely non-stock — exclude from match pool
            non_stock_count += 1
            non_stock_total += net_non_stock
            continue

        if net_stock == 0 and net_non_stock == 0:
            # Zero-value invoice — nothing to reconcile
            continue

        # Build line item summaries for insight generation
        line_summaries = []
        if inv.line_items:
            for item in inv.line_items:
                line_summaries.append({
                    "description": item.description or "",
                    "amount": str(item.amount or 0),
                    "is_non_stock": item.is_non_stock,
                })

        stock_invoices.append({
            "id": inv.id,
            "date": inv_date,
            "supplier_name": supplier_name,
            "ref": inv.invoice_number or "",
            "net_stock": net_stock.quantize(Decimal("0.01")),
            "net_non_stock": net_non_stock.quantize(Decimal("0.01")),
            "document_type": inv.document_type or "invoice",
            "line_items": line_summaries,
        })

        # Track non-stock portion of mixed invoices separately
        if net_non_stock != 0:
            non_stock_count += 1
            non_stock_total += net_non_stock

    return stock_invoices, non_stock_count, non_stock_total.quantize(Decimal("0.01"))


# ============ Build supplier alias map from DB ============

async def build_supplier_alias_map(db: AsyncSession, kitchen_id: int) -> dict[str, str]:
    """
    Build a normalised-name → canonical-name map from the Supplier table aliases.
    Merges with the hardcoded SUPPLIER_ALIASES.
    """
    alias_map = dict(SUPPLIER_ALIASES)  # start with hardcoded

    result = await db.execute(
        select(Supplier).where(Supplier.kitchen_id == kitchen_id)
    )
    suppliers = result.scalars().all()

    for s in suppliers:
        # Map canonical name
        alias_map[normalise_supplier(s.name)] = s.name
        # Map aliases
        if s.aliases:
            for alias in s.aliases:
                alias_map[normalise_supplier(alias)] = s.name

    return alias_map


def resolve_supplier_with_map(xero_desc: str, alias_map: dict[str, str]) -> str:
    """Resolve a Xero description to a canonical supplier name using dynamic alias map."""
    extracted = extract_supplier_from_xero_desc(xero_desc)
    norm = normalise_supplier(extracted)
    if norm in alias_map:
        return alias_map[norm]
    # Try partial match — check if any alias key is contained in the extracted name
    for alias_key, canonical in alias_map.items():
        if alias_key in norm or norm in alias_key:
            return canonical
    return extracted


def suppliers_match(flash_supplier: str, xero_supplier: str) -> bool:
    """Check if a Flash supplier and resolved Xero supplier match."""
    return normalise_supplier(flash_supplier) == normalise_supplier(xero_supplier)


# ============ Matching Engine ============

def run_matching(
    flash_items: list[dict],
    xero_items: list[dict],
    alias_map: dict[str, str],
) -> tuple[list, list, list[dict], list[dict]]:
    """
    Three-pass matching:
    1. Exact: same normalised ref + same date + amount within £0.01
    2. Near-match: one field differs (flagged as discrepancy)
    3. Leftovers: unmatched on each side

    Returns (matched, discrepancies, flash_unmatched, xero_unmatched)
    """
    matched = []
    discrepancies = []

    flash_used = set()
    xero_used = set()

    # Resolve Xero suppliers up front
    for xi, xero in enumerate(xero_items):
        xero["_resolved_supplier"] = resolve_supplier_with_map(xero["description"], alias_map)
        xero["_norm_ref"] = normalise_ref(xero["ref"])
        xero["_idx"] = xi

    for fi, flash in enumerate(flash_items):
        flash["_norm_ref"] = normalise_ref(flash["ref"])
        flash["_idx"] = fi

    # ---- Pass 1: Exact match ----
    for fi, flash in enumerate(flash_items):
        if fi in flash_used:
            continue
        if not flash["_norm_ref"]:
            continue  # Can't exact-match without a ref
        for xi, xero in enumerate(xero_items):
            if xi in xero_used:
                continue
            if flash["_norm_ref"] == xero["_norm_ref"] and \
               flash["date"] == xero["date"] and \
               amounts_match(flash["net_stock"], xero["net"]):
                matched.append({
                    "date": flash["date"].isoformat(),
                    "supplier": flash["supplier_name"],
                    "ref": flash["ref"] or xero["ref"],
                    "amount": fmt_money(flash["net_stock"]),
                    "flash_id": flash["id"],
                })
                flash_used.add(fi)
                xero_used.add(xi)
                break

    # ---- Pass 2: Near-match (discrepancies) ----
    for fi, flash in enumerate(flash_items):
        if fi in flash_used:
            continue
        best_match = None
        best_score = 0  # higher = more confident

        for xi, xero in enumerate(xero_items):
            if xi in xero_used:
                continue

            differs = []
            score = 0

            f_ref = flash["_norm_ref"]
            x_ref = xero["_norm_ref"]
            f_supplier = flash["supplier_name"]
            x_supplier = xero["_resolved_supplier"]
            same_supplier = suppliers_match(f_supplier, x_supplier)

            # 2a: Same ref, same date, amount differs
            if f_ref and x_ref and f_ref == x_ref and flash["date"] == xero["date"]:
                if not amounts_match(flash["net_stock"], xero["net"]):
                    differs.append("amount")
                    score = 10

            # 2b: Same ref, amount matches, date differs ≤ 3 days
            elif f_ref and x_ref and f_ref == x_ref and \
                 amounts_match(flash["net_stock"], xero["net"]):
                date_diff = abs((flash["date"] - xero["date"]).days)
                if date_diff <= 3:
                    differs.append("date")
                    score = 9

            # 2c: Same supplier, amount matches (within £0.01), date within 3 days, refs differ
            # This catches cases where refs are completely different
            # (e.g. Cotswold Coffee 126621 vs 145323, R&D Walker Nº4 vs 106891)
            elif same_supplier and amounts_match(flash["net_stock"], xero["net"], Decimal("0.01")):
                date_diff = abs((flash["date"] - xero["date"]).days)
                if date_diff == 0:
                    differs.append("ref")
                    score = 8
                elif date_diff <= 3:
                    differs.extend(["ref", "date"])
                    score = 7

            # 2d: Same supplier, same ref, both amount and date differ slightly
            elif f_ref and x_ref and f_ref == x_ref and same_supplier:
                date_diff = abs((flash["date"] - xero["date"]).days)
                amt_diff = abs(flash["net_stock"] - xero["net"])
                if date_diff <= 3 and amt_diff <= Decimal("5.00"):
                    differs.extend(["amount", "date"])
                    score = 5

            # 2e: Credit note matching — same supplier, same amount,
            # but refs differ completely (Xero uses CN prefix, Flash uses supplier ref)
            elif same_supplier and amounts_match(flash["net_stock"], xero["net"]) and \
                 (flash["document_type"] == "credit_note" or xero.get("is_credit_note", False)):
                date_diff = abs((flash["date"] - xero["date"]).days)
                if date_diff <= 5:
                    differs.append("ref")
                    if date_diff > 0:
                        differs.append("date")
                    score = 8  # High confidence — credit notes with matching supplier+amount

            # 2f: Same supplier, date within 3 days, amounts close but not exact
            elif same_supplier:
                date_diff = abs((flash["date"] - xero["date"]).days)
                amt_diff = abs(flash["net_stock"] - xero["net"])
                if date_diff <= 3 and amt_diff <= Decimal("5.00") and amt_diff > Decimal("0.01"):
                    differs_list = ["amount"]
                    if date_diff > 0:
                        differs_list.append("date")
                    if f_ref != x_ref:
                        differs_list.append("ref")
                    differs.extend(differs_list)
                    score = 4

            if score > best_score:
                best_score = score
                best_match = (xi, xero, differs)

        if best_match:
            xi, xero, differs = best_match
            f_net = flash["net_stock"]
            x_net = xero["net"]
            diff_val = f_net - x_net
            amt_diff_abs = abs(float(diff_val))

            # Safety check: reject any near-match where amount differs by more than £5
            # AND refs are different — these are almost certainly different invoices
            f_ref_norm = normalise_ref(flash["ref"])
            x_ref_norm = normalise_ref(xero["ref"])
            refs_differ = f_ref_norm != x_ref_norm

            if refs_differ and amt_diff_abs > 5.0:
                logger.warning(
                    f"REJECTED false discrepancy: Flash {flash['supplier_name']} "
                    f"{flash['ref']}={f_net} vs Xero {xero['ref']}={x_net} "
                    f"diff={amt_diff_abs:.2f} score={best_score}"
                )
                continue  # Skip — leave both as unmatched

            # Add "amount" to differs if amounts don't actually match
            if amt_diff_abs > 0.01 and "amount" not in differs:
                differs.append("amount")

            logger.info(
                f"Discrepancy: Flash {flash['supplier_name']} {flash['ref']}={f_net} "
                f"vs Xero {xero['ref']}={x_net} differs={differs} score={best_score}"
            )

            discrepancies.append({
                "date": flash["date"].isoformat(),
                "supplier": flash["supplier_name"],
                "ref": flash["ref"] or xero["ref"],
                "flash_amount": fmt_money(f_net),
                "xero_amount": fmt_money(x_net),
                "difference": fmt_money(diff_val),
                "differs": differs,
                "flash_id": flash["id"],
                "flash_date": flash["date"].isoformat(),
                "xero_date": xero["date"].isoformat(),
                "flash_ref": flash["ref"],
                "xero_ref": xero["ref"],
                "flash_supplier": flash["supplier_name"],
                "xero_description": xero["description"],
            })
            flash_used.add(fi)
            xero_used.add(xi)

    # ---- Leftovers ----
    flash_unmatched = [f for fi, f in enumerate(flash_items) if fi not in flash_used]
    xero_unmatched = [x for xi, x in enumerate(xero_items) if xi not in xero_used]

    return matched, discrepancies, flash_unmatched, xero_unmatched


# ============ Amount Discrepancy Insights ============

def check_non_stock_explains_diff(flash: dict, diff_val: Decimal) -> str | None:
    """
    Check if the amount discrepancy is explained by non-stock items.
    If Xero includes the full invoice (stock + non-stock) but Flash only shows stock,
    the difference should equal the non-stock total.
    """
    net_non_stock = flash.get("net_non_stock", Decimal("0"))
    if net_non_stock == 0:
        return None

    # diff_val = flash_stock - xero_net (negative when Xero is higher)
    # If Xero has full invoice, diff = -net_non_stock
    if amounts_match(abs(diff_val), abs(net_non_stock), Decimal("0.02")):
        non_stock_items = [
            li for li in flash.get("line_items", []) if li.get("is_non_stock")
        ]
        item_names = ", ".join(
            li["description"][:40] for li in non_stock_items if li.get("description")
        )
        return (
            f"Non-stock items account for the difference "
            f"({fmt_money(abs(net_non_stock))}). "
            f"Xero likely includes full invoice total. "
            f"Non-stock: {item_names}" if item_names else
            f"Non-stock items account for the difference "
            f"({fmt_money(abs(net_non_stock))}). "
            f"Xero likely includes full invoice total."
        )

    # Check if non-stock is a partial explanation (diff is larger but non-stock is a chunk)
    if abs(net_non_stock) > Decimal("1.00") and abs(diff_val) > abs(net_non_stock):
        remainder = abs(diff_val) - abs(net_non_stock)
        return (
            f"Non-stock items total {fmt_money(abs(net_non_stock))} "
            f"but difference is {fmt_money(abs(diff_val))} — "
            f"non-stock explains part, {fmt_money(remainder)} remains unexplained."
        )

    return None


def check_line_item_combinations(flash: dict, diff_val: Decimal) -> str | None:
    """
    Check if any single line item or small combination matches the difference.
    This catches cases where a specific item was excluded/included differently.
    """
    line_items = flash.get("line_items", [])
    if not line_items:
        return None

    target = abs(diff_val)

    # Check single items
    for li in line_items:
        amt = abs(Decimal(li["amount"]))
        if amounts_match(amt, target, Decimal("0.02")) and amt > Decimal("0.50"):
            desc = li["description"][:50] if li["description"] else "unnamed item"
            ns = " (non-stock)" if li.get("is_non_stock") else ""
            return f'Single line item matches difference: "{desc}"{ns} = {fmt_money(amt)}'

    return None


async def generate_llm_insight(
    db: AsyncSession,
    kitchen_id: int,
    flash: dict,
    xero_net: Decimal,
    diff_val: Decimal,
) -> str | None:
    """Use LLM to analyse line items and suggest cause of amount variance."""
    from services.llm_service import call_llm

    line_items = flash.get("line_items", [])
    if not line_items:
        return None

    # Build concise line item list
    li_summary = []
    for li in line_items:
        li_summary.append({
            "description": li["description"][:60] if li["description"] else "—",
            "amount": li["amount"],
            "non_stock": li["is_non_stock"],
        })

    system_msg = (
        "You are a kitchen accounts assistant. Flash is a food stock/GP system that "
        "tracks invoices — it separates stock (food) items from non-stock items "
        "(chemicals, packaging, equipment, etc). Xero is the accounting system that "
        "records the full invoice total posted to the food purchases account. "
        "An amount discrepancy means Flash stock total differs from the Xero net. "
        "Common causes: non-stock items not split out in Xero, line items missing, "
        "rounding, or Xero posting error."
    )

    user_msg = (
        f"Invoice: {flash['supplier_name']} ref {flash['ref']} dated {flash['date']}\n"
        f"Flash stock total: {fmt_money(flash['net_stock'])}\n"
        f"Xero net: {fmt_money(xero_net)}\n"
        f"Difference: {fmt_money(diff_val)} (Flash - Xero)\n\n"
        f"Flash line items:\n{json.dumps(li_summary, indent=2)}\n\n"
        f"Can you identify which line items or combination likely accounts for "
        f"the {fmt_money(abs(diff_val))} difference? "
        f"Reply in ONE short sentence (max 120 chars). "
        f"If unclear, say 'Unable to determine cause'."
    )

    result = await call_llm(
        db=db,
        kitchen_id=kitchen_id,
        feature="reconciliation_insight",
        messages=[{"role": "user", "content": user_msg}],
        system_message=system_msg,
    )

    if result["status"] not in ("success", "cached"):
        return None

    text = result.get("result", "")
    if isinstance(text, dict):
        text = str(text)
    text = text.strip().strip('"').strip("'")
    if text and len(text) < 200:
        return text
    return text[:200] + "..." if text else None


async def generate_amount_insights(
    db: AsyncSession,
    kitchen_id: int,
    discrepancies: list[dict],
    flash_lookup: dict[int, dict],
) -> list[dict]:
    """
    For each amount discrepancy, try to explain the variance:
    1. Code check: does non-stock total match the difference?
    2. Code check: does a single line item match the difference?
    3. LLM fallback: send line items for analysis
    """
    for disc in discrepancies:
        if "amount" not in disc.get("differs", []):
            continue

        flash_id = disc.get("flash_id")
        if not flash_id or flash_id not in flash_lookup:
            continue

        flash = flash_lookup[flash_id]

        # Parse the difference back to Decimal
        diff_str = disc["difference"].replace("£", "").replace(",", "")
        try:
            diff_val = Decimal(diff_str)
        except InvalidOperation:
            continue

        # 1. Non-stock check
        insight = check_non_stock_explains_diff(flash, diff_val)
        if insight:
            disc["amount_insight"] = insight
            continue

        # 2. Single line item check
        insight = check_line_item_combinations(flash, diff_val)
        if insight:
            disc["amount_insight"] = insight
            continue

        # 3. LLM fallback
        xero_str = disc["xero_amount"].replace("£", "").replace(",", "")
        try:
            xero_net = Decimal(xero_str)
        except InvalidOperation:
            continue

        insight = await generate_llm_insight(db, kitchen_id, flash, xero_net, diff_val)
        if insight:
            disc["amount_insight"] = f"🤖 {insight}"

    return discrepancies


# ============ LLM Fallback Matching ============

async def llm_match_remaining(
    db: AsyncSession,
    kitchen_id: int,
    flash_unmatched: list[dict],
    xero_unmatched: list[dict],
) -> tuple[list, list, list[dict], list[dict]]:
    """
    Use LLM to attempt matching remaining unmatched items.
    Returns (new_discrepancies, new_exact, remaining_flash, remaining_xero)
    """
    from services.llm_service import call_llm

    if not flash_unmatched or not xero_unmatched:
        return [], [], flash_unmatched, xero_unmatched

    # Build concise representations
    flash_summary = []
    for i, f in enumerate(flash_unmatched):
        flash_summary.append({
            "idx": i,
            "date": f["date"].isoformat(),
            "supplier": f["supplier_name"],
            "ref": f["ref"],
            "amount": str(f["net_stock"]),
        })

    xero_summary = []
    for i, x in enumerate(xero_unmatched):
        xero_summary.append({
            "idx": i,
            "date": x["date"].isoformat(),
            "description": x["description"],
            "ref": x["ref"],
            "amount": str(x["net"]),
        })

    system_msg = """You are a bookkeeping reconciliation assistant. You are given two lists of unmatched invoice entries — one from Flash (the kitchen invoice system) and one from Xero (the accounting system).

Your job is to identify probable matches between the two lists. These are entries that likely represent the same real-world invoice but have data discrepancies (different reference numbers, slightly different amounts, date offsets, supplier name variants, etc).

For each probable match, explain which fields differ and why you think they are the same invoice.

IMPORTANT: Only suggest matches you are reasonably confident about. It is better to leave items unmatched than to create false matches. Consider supplier names, dates, amounts, and reference numbers holistically."""

    user_msg = f"""Here are the unmatched Flash invoices:
{json.dumps(flash_summary, indent=2)}

Here are the unmatched Xero entries:
{json.dumps(xero_summary, indent=2)}

Return a JSON array of matches. Each match should be:
{{
  "flash_idx": <index in flash list>,
  "xero_idx": <index in xero list>,
  "confidence": "high" or "medium",
  "differs": ["field1", "field2"],
  "reasoning": "brief explanation"
}}

Only include matches with high or medium confidence. Return an empty array [] if no good matches found."""

    result = await call_llm(
        db=db,
        kitchen_id=kitchen_id,
        feature="reconciliation_matching",
        messages=[{"role": "user", "content": user_msg}],
        system_message=system_msg,
    )

    if result["status"] not in ("success", "cached"):
        logger.info(f"LLM reconciliation matching unavailable: {result['status']}")
        return [], [], flash_unmatched, xero_unmatched

    # Parse LLM response
    new_discrepancies = []
    llm_text = result.get("result", "")
    if isinstance(llm_text, dict):
        llm_text = json.dumps(llm_text)
    if not llm_text:
        return [], [], flash_unmatched, xero_unmatched

    try:
        # Extract JSON from response (may be wrapped in markdown code block)
        json_match = re.search(r"\[.*\]", str(llm_text), re.DOTALL)
        if not json_match:
            return [], [], flash_unmatched, xero_unmatched
        matches = json.loads(json_match.group())
    except (json.JSONDecodeError, AttributeError):
        logger.warning("Failed to parse LLM reconciliation response")
        return [], [], flash_unmatched, xero_unmatched

    flash_matched = set()
    xero_matched = set()

    for m in matches:
        fi = m.get("flash_idx")
        xi = m.get("xero_idx")
        confidence = m.get("confidence", "medium")

        if fi is None or xi is None:
            continue
        if fi >= len(flash_unmatched) or xi >= len(xero_unmatched):
            continue
        if fi in flash_matched or xi in xero_matched:
            continue

        flash = flash_unmatched[fi]
        xero = xero_unmatched[xi]

        # Safety check: reject LLM matches where refs differ AND amount diff > £5
        f_ref_norm = normalise_ref(flash["ref"])
        x_ref_norm = normalise_ref(xero["ref"])
        amt_diff_abs = abs(float(flash["net_stock"] - xero["net"]))
        if f_ref_norm != x_ref_norm and amt_diff_abs > 5.0:
            logger.warning(
                f"REJECTED LLM false match: Flash {flash['supplier_name']} "
                f"{flash['ref']}={flash['net_stock']} vs Xero {xero['ref']}={xero['net']} "
                f"diff={amt_diff_abs:.2f}"
            )
            continue

        differs = m.get("differs", [])
        if not differs:
            differs = ["unknown"]

        diff_val = flash["net_stock"] - xero["net"]
        new_discrepancies.append({
            "date": flash["date"].isoformat(),
            "supplier": flash["supplier_name"],
            "ref": flash["ref"] or xero["ref"],
            "flash_amount": fmt_money(flash["net_stock"]),
            "xero_amount": fmt_money(xero["net"]),
            "difference": fmt_money(diff_val),
            "differs": differs,
            "flash_id": flash["id"],
            "flash_date": flash["date"].isoformat(),
            "xero_date": xero["date"].isoformat(),
            "flash_ref": flash["ref"],
            "xero_ref": xero["ref"],
            "match_source": "ai",
        })
        flash_matched.add(fi)
        xero_matched.add(xi)

    remaining_flash = [f for i, f in enumerate(flash_unmatched) if i not in flash_matched]
    remaining_xero = [x for i, x in enumerate(xero_unmatched) if i not in xero_matched]

    return new_discrepancies, [], remaining_flash, remaining_xero


# ============ Main Endpoint ============

@router.post("/purchases/reconcile", response_model=ReconcileResponse)
async def reconcile_purchases(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Upload a Xero Account Transactions XLSX and reconcile against Flash invoices.
    """
    # Validate file type
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an XLSX file")

    # Read file into memory (no persistence)
    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")

    # Parse XLSX
    try:
        period_start, period_end, xero_rows = parse_xero_xlsx(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"XLSX parsing failed: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {str(e)}")

    # Group Xero rows by reference
    xero_items = group_xero_by_ref(xero_rows)

    # Get Flash invoices for the period
    flash_items, non_stock_count, non_stock_total = await get_flash_invoices(
        db, current_user.kitchen_id, period_start, period_end
    )

    # Build supplier alias map from DB
    alias_map = await build_supplier_alias_map(db, current_user.kitchen_id)

    # Run rule-based matching
    matched, discrepancies, flash_unmatched, xero_unmatched = run_matching(
        flash_items, xero_items, alias_map
    )

    # LLM fallback for remaining unmatched
    llm_attempted = False
    if flash_unmatched and xero_unmatched:
        llm_disc, llm_exact, flash_unmatched, xero_unmatched = await llm_match_remaining(
            db, current_user.kitchen_id, flash_unmatched, xero_unmatched
        )
        if llm_disc or llm_exact:
            llm_attempted = True
            discrepancies.extend(llm_disc)
            matched.extend(llm_exact)

    # Generate insights for amount discrepancies
    flash_lookup = {f["id"]: f for f in flash_items}
    discrepancies = await generate_amount_insights(
        db, current_user.kitchen_id, discrepancies, flash_lookup
    )

    # Calculate totals
    flash_total = sum(f["net_stock"] for f in flash_items)
    xero_total = sum(x["net"] for x in xero_items)
    difference = flash_total - xero_total

    # Build response
    matched_response = [ReconcileMatchedItem(**m) for m in matched]

    discrepancy_response = [ReconcileDiscrepancyItem(**d) for d in discrepancies]

    flash_only_response = [
        ReconcileUnmatchedFlash(
            date=f["date"].isoformat(),
            supplier=f["supplier_name"],
            ref=f["ref"],
            net_stock=fmt_money(f["net_stock"]),
            flash_id=f["id"],
        )
        for f in flash_unmatched
    ]

    xero_only_response = [
        ReconcileUnmatchedXero(
            date=x["date"].isoformat(),
            description=x["description"],
            ref=x["ref"],
            net=fmt_money(x["net"]),
            is_expected_external=is_expected_external(x["description"]),
        )
        for x in xero_unmatched
    ]

    return ReconcileResponse(
        period_start=period_start.isoformat(),
        period_end=period_end.isoformat(),
        flash_total=fmt_money(flash_total),
        xero_total=fmt_money(xero_total),
        difference=fmt_money(difference),
        matched_count=len(matched_response),
        discrepancy_count=len(discrepancy_response),
        flash_only_count=len(flash_only_response),
        xero_only_count=len(xero_only_response),
        non_stock_excluded_count=non_stock_count,
        non_stock_excluded_total=fmt_money(non_stock_total),
        matched=matched_response,
        discrepancies=discrepancy_response,
        flash_only=flash_only_response,
        xero_only=xero_only_response,
        llm_matches_attempted=llm_attempted,
    )
